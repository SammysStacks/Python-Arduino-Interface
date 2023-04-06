#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Mar  2 13:56:47 2021

@author: samuelsolomon
"""

# -------------------------------------------------------------------------- #
# ---------------------------- Imported Modules ---------------------------- #

# General Modules
import os
import sys
import numpy as np
import pandas as pd
import time as timer
from scipy import stats
# Module to Sort Files in Order
from natsort import natsorted
# Read/Write to Excel
import csv
import pyexcel
import openpyxl as xl
from openpyxl import load_workbook, Workbook
# Openpyxl Styles
from openpyxl.styles import Alignment
from openpyxl.styles import Font

# -------------------------------------------------------------------------- #
# ---------------------- Extract Test Data from Excel ---------------------- #

class handlingExcelFormat:   

    def __init__(self):
        # Hardcoded sheetnames for different types of excel information
        self.emptySheetName = "empty"
        self.rawSignals_Sheetname = "Raw Signals; File 0"
        self.subjectInfo_SheetName = "Subject Info; File 0"
        self.rawFeatures_AppendedSheetName = " Features; File 0"
        self.experimentalInfo_SheetName = "Experimental Info; File 0"
        
        # Hardcoded folder names
        self.saveFeatureFolder = "Saved Features/"
        
        # Hardcoded file names
        self.saveFeatureFile_Appended = " Features.xlsx"
        
        # Excel parameters
        self.maxAddToexcelSheet = 1048500  # Max Rows in a Worksheet
        
    def convertToXLSX(self, inputExcelFile):
        """
        Converts .xls Files to .xlsx Files That OpenPyxl Can Read
        If the File is Already a .xlsx Files, Do Nothing
        If the File is Neither a .xls Nor .xlsx, it Exits the Program
        """
        # Check That the Current Extension is .xls or .xlsx
        _, extension = os.path.splitext(inputExcelFile)
        # If the Extension is .xlsx, the File is Ready; Do Nothing
        if extension == '.xlsx':
            return inputExcelFile
        # If the Extension is Not .xls/.xlsx, Then the Data is in the Wrong Format; Exit Program
        if extension not in ['.xls', '.xlsx']:
            print("Cannot Convert File to .xlsx")
            sys.exit()
        
        # Create Output File Directory to Save Data ONLY If None Exists
        newExcelFolder = os.path.dirname(inputExcelFile) + "/Excel Files/"
        os.makedirs(newExcelFolder, exist_ok = True)
        
        # Convert '.xls' to '.xlsx'
        filename = os.path.basename(inputExcelFile)
        newExcelFile = newExcelFolder + filename + "x"
        pyexcel.save_as(file_name = inputExcelFile, dest_file_name = newExcelFile, logfile=open(os.devnull, 'w'))
        
        # Save New Excel name
        return newExcelFile
    
    def txt2csv(self, txtFile, csvFile, csvDelimiter = ",", overwriteCSV = False):
        # Check to see if csv conversion alreayd happened
        if not os.path.isfile(csvFile) or overwriteCSV:
            with open(txtFile, "r") as inputData:
                in_reader = csv.reader(inputData, delimiter = csvDelimiter)
                with open(csvFile, 'w', newline='') as out_csv:
                    out_writer = csv.writer(out_csv)
                    for row in in_reader:
                        out_writer.writerow(row)
    
    def convertToExcel(self, inputFile, excelFile, excelDelimiter = ",", overwriteXL = False, testSheetNum = 0):
        # If the File is Not Already Converted: Convert the CSV to XLSX
        if not os.path.isfile(excelFile) or overwriteXL:
            if excelDelimiter == "fixedWidth":
                df = pd.read_fwf(inputFile)
                df.drop(index=0, inplace=True) # drop the underlines
                df.to_excel(excelFile, index=False)
                # Load the Data from the Excel File
                xlWorkbook = xl.load_workbook(excelFile, data_only=True, read_only=True)
                xlWorksheets = xlWorkbook.worksheets[testSheetNum:]
            else:
                # Make Excel WorkBook
                xlWorkbook = xl.Workbook()
                xlWorksheet = xlWorkbook.active
                # Write the Data from the CSV File to the Excel WorkBook
                with open(inputFile, "r") as inputData:
                    inReader = csv.reader(inputData, delimiter = excelDelimiter)
                    with open(excelFile, 'w+', newline=''):
                        for row in inReader:
                            xlWorksheet.append(row)    
                # Save as New Excel File
                xlWorkbook.save(excelFile)
                xlWorksheets = [xlWorksheet]
        # Else Load the Data from the Excel File
        else:
            # Load the Data from the Excel File
            xlWorkbook = xl.load_workbook(excelFile, data_only=True, read_only=True)
            xlWorksheets = xlWorkbook.worksheets[testSheetNum:]
        
        # Return Excel Sheet
        return xlWorkbook, xlWorksheets
    
    def splitExcelSheetsToExcelFiles(self, inputFile):
        wb = load_workbook(filename=inputFile)
        
        for sheet in wb.worksheets:
            new_wb = Workbook()
            ws = new_wb.active
            for row_data in sheet.iter_rows():
                for row_cell in row_data:
                    ws[row_cell.coordinate].value = row_cell.value
        
            new_wb.save('{0}.xlsx'.format(sheet.title))
    
    def addExcelAesthetics(self, worksheet):
        # Initialize variables
        align = Alignment(horizontal='center',vertical='center',wrap_text=True) 
        
        # Loop through each header cell
        for headerCell in worksheet[1]:
            column_cells = worksheet[headerCell.column_letter]
            
            # Set the column width
            length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
            worksheet.column_dimensions[headerCell.column_letter].width = max(length, worksheet.column_dimensions[headerCell.column_letter].width)
            worksheet.column_dimensions[headerCell.column_letter].bestFit = True
            # Center the Data in the Cells
            for cell in column_cells:
                cell.alignment = align
            # Set the header text color
            headerCell.font = Font(color='00FF0000', italic=True, bold=True)
        
        return worksheet
    
    
class getExcelData(handlingExcelFormat):
    
    def extractFeatureNames(self, featureLabelFile, prependedString, appendToName = ''):
        """ Extract the Feature Names from a txt File """
        # Check if File Exists
        if not os.path.exists(featureLabelFile):
            print("The following Input File Does Not Exist:", featureLabelFile)
            sys.exit()

        # Get the Data
        fullText = ''
        with open(featureLabelFile, "r", newline='\n') as inputData:
            inReader = csv.reader(inputData)
            for row in inReader:
                for featureString in row:
                    if featureString[0] != "#":
                        fullText += featureString + ","
        
        possibleFeatures = fullText.split(prependedString)
        # Extract the Features
        featureList = []
        for feature in possibleFeatures:
            feature = feature.split("[")[-1]
            feature = feature.split("]")[0]
            feature = feature.replace(" ", "")
            feature = feature.replace("\n", "")
            
            if len(feature) != 0:
                feature = feature.split(",")
                featureList.extend(feature)
                
        featureListFull = []
        for feature in featureList:
            featureListFull.append(feature + appendToName)
        
        return featureListFull
    
    def extractRawSignalData(self, excelSheet, startDataCol = 1, endDataCol = 2, data = None):
        # If Header Exists, Skip Until You Find the Data
        for row in excelSheet.rows:
            cellA = row[0]
            if type(cellA.value) in [int, float]:
                dataStartRow = cellA.row + 1
                break
        
        if data == None:
            data = [ [], [[] for channel in range(endDataCol-startDataCol)] ]
        # Loop Through the Excel Worksheet to collect all the data
        for dataRow in excelSheet.iter_rows(min_col=startDataCol, min_row=dataStartRow-1, max_col=endDataCol, max_row=excelSheet.max_row):
            # Stop Collecting Data When there is No More
            if dataRow[0].value == None:
                break
            
            # Get Data
            data[0].append(float(dataRow[0].value))
            for dataInd in range(1, len(dataRow)):
                data[1][dataInd-1].append(float(dataRow[dataInd].value or 0))
        
        return data
    
    def extractExperimentalInfo(self, excelSheet, experimentTimes = [], experimentNames = [], surveyAnswerTimes = [], surveyAnswersList = [], surveyQuestions = []):
        # If Header Exists, Skip Until You Find the Data
        for row in excelSheet.rows:
            cellA = row[0]
            if type(cellA.value) in [int, float]:
                dataStartRow = cellA.row + 1
                endDataCol = len(row)
                break
            elif type(cellA.value) == str:
                # Extract the survey questions if none given
                if len(surveyQuestions) == 0:
                    surveyQuestions = np.array([str(cell.value) for cell in row[4:]], dtype=str)
                else:
                    # Assert the survey questions are the same in all instances.
                    assert all(surveyQuestions == np.array([str(cell.value) for cell in row[4:]], dtype=str)), "We have two experimental info sheets with DIFFERENT features"
        
        # Loop Through the Excel Worksheet to collect all the data
        for dataRow in excelSheet.iter_rows(min_col=1, min_row=dataStartRow-1, max_col=4, max_row=excelSheet.max_row):
            # Stop Collecting Data When there is No More
            if dataRow[0].value == None:
                break
            
            # Get Data
            endExperimentTime = float(dataRow[1].value) if dataRow[1].value != None else dataRow[1].value
            experimentTimes.append([float(dataRow[0].value), endExperimentTime])
            experimentNames.append(str(dataRow[2].value))
        
        # Loop Through the Excel Worksheet to collect all the data
        for dataRow in excelSheet.iter_rows(min_col=4, min_row=dataStartRow-1, max_col=endDataCol, max_row=excelSheet.max_row):
            # Stop Collecting Data When there is No More
            if dataRow[0].value == None:
                break
            
            # Get Data
            surveyAnswerTimes.append(float(dataRow[0].value))
            surveyAnswersList.append([float(dataRow[dataInd].value or 0) for dataInd in range(1, len(dataRow))])

        
        return experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions
    
    def extractSubjectInfo(self, excelSheet, subjectInformationAnswers = [], subjectInformationQuestions = []):        
        # If Header Exists, Skip Until You Find the Data
        for row in excelSheet.rows:
            cellA = row[0]
            if type(cellA.value) == str:
                dataStartRow = cellA.row + 1
                break
            
        # Loop Through the Excel Worksheet to collect all the data
        for dataRow in excelSheet.iter_rows(min_col=1, min_row=dataStartRow, max_col=2, max_row=excelSheet.max_row):
            # Stop Collecting Data When there is No More
            if dataRow[0].value == None:
                break
            
            # Get Data
            subjectInformationAnswers.append(str(dataRow[1].value))
            subjectInformationQuestions.append(str(dataRow[0].value))
        
        return subjectInformationAnswers, subjectInformationQuestions

    def extractExperimentalData(self, worksheets, numberOfChannels, surveyQuestions = [], finalSubjectInformationQuestions = []):
        # Initialize data holder
        compiledRawData = [ [], [[] for channel in range(numberOfChannels)] ]
        # Initialize experimental information
        experimentTimes = []; experimentNames = []
        surveyAnswerTimes = []; surveyAnswersList = [];
        # Initialize suject information
        subjectInformationAnswers = []; subjectInformationQuestions = []
        
        # Loop through and compile all the data in the file
        for excelSheet in worksheets:
            # Extract experiment information
            if self.experimentalInfo_SheetName in excelSheet.title:
                experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions = self.extractExperimentalInfo(excelSheet, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions)
            # Extract subject information
            elif self.subjectInfo_SheetName in excelSheet.title:
                subjectInformationAnswers, subjectInformationQuestions = self.extractSubjectInfo(excelSheet, subjectInformationAnswers, subjectInformationQuestions)
            # Extract Time and Current Data from the File
            else:
                compiledRawData = self.extractRawSignalData(excelSheet, startDataCol = 1, endDataCol = 1 + numberOfChannels, data = compiledRawData)
        
        # Check the data integrity
        if len(compiledRawData[0]) == 0:
            print("\tNo data found in this file")
        # Check that the subject background questions are all the same
        if len(finalSubjectInformationQuestions) != 0:
            assert all(np.array(finalSubjectInformationQuestions) == subjectInformationQuestions)
        
        return compiledRawData, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions, subjectInformationAnswers, subjectInformationQuestions
    
    def getData(self, inputFile, numberOfChannels = 1, testSheetNum = 0):
        """
        Extracts Pulse Data from Excel Document (.xlsx). Data can be in any
        worksheet which the user can specify using 'testSheetNum' (0-indexed).
        In the Worksheet:
            Time Data must be in Column 'A' (x-Axis)
            Biolectric Data must be in Column 'B-x' (y-Axis)
        If No Data is present in one cell of a row, it will be read in as zero.
        --------------------------------------------------------------------------
        Input Variable Definitions:
            inputFile: The Path to the Excel/TXT/CSV File Containing the Biolectric Data.
            numberOfChannels: The number of biolectric signals to extract.
            testSheetNum: An Integer Representing the Excel Worksheet (0-indexed) to Begin on.
        --------------------------------------------------------------------------
        """
        # Check if File Exists
        if not os.path.exists(inputFile):
            print("The following Input File Does Not Exist:", inputFile)
            sys.exit()
            
        # Convert to TXT and CSV Files to XLSX
        if inputFile.endswith(".txt") or inputFile.endswith(".csv"):
            # Extract Filename Information
            oldFileExtension = os.path.basename(inputFile)
            filename = os.path.splitext(oldFileExtension)[0]
            newFilePath = os.path.dirname(inputFile) + "/Excel Files/"
            # Make Output Folder Directory if Not Already Created
            os.makedirs(newFilePath, exist_ok = True)

            # Convert CSV or TXT to XLSX
            excelFile = newFilePath + filename + ".xlsx"
            xlWorkbook, worksheets = self.convertToExcel(inputFile, excelFile, excelDelimiter = ",", overwriteXL = False, testSheetNum = testSheetNum)
        # If the File is Already an Excel File, Just Load the File
        elif inputFile.endswith(".xlsx"):
            # Load the Data from the Excel File
            xlWorkbook = xl.load_workbook(inputFile, data_only=True, read_only=True)
            worksheets = xlWorkbook.worksheets[testSheetNum:]
        else:
            print("The Following File is Neither CSV, TXT, Nor XLSX:", inputFile)
        print("Extracting Data from the Excel File:", inputFile)
        
        # Extract the data
        compiledRawData, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions, subjectInformationAnswers, subjectInformationQuestions = self.extractExperimentalData(worksheets, numberOfChannels)
        xlWorkbook.close()
        
        # Finished Data Collection: Close Workbook and Return Data to User
        print("\tFinished Collecting Biolectric Data");
        return compiledRawData, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions, subjectInformationAnswers, subjectInformationQuestions

    def extractFeatures(self, excelSheet, featureOrder, features, featuresTimesHolder, featureNames):            
        # Find the type of features we are extracting
        featureType = excelSheet.title.split(" ")[0].lower()
        featureInd = featureOrder.index(featureType)
        
        # If Header Exists, Skip Until You Find the Data
        for row in excelSheet.rows:
            cellA = row[0]
            if type(cellA.value) in [int, float]:
                dataStartRow = cellA.row + 1
                endDataCol = len(row)
                break
            elif type(cellA.value) == str:
                # If no feature names found, save them
                if len(featureNames[featureInd]) == 0:
                    featureNames[featureInd] = np.array([str(cell.value) for cell in row[1:]], dtype = str)
                else:
                    # Assert the same feature names present in all files.
                    assert all(featureNames[featureInd] == np.array([str(cell.value) for cell in row[1:]], dtype = str)), "We have two feature sheets with DIFFERENT features for " + featureType
            
        # Loop Through the Excel Worksheet to collect all the data
        for dataRow in excelSheet.iter_rows(min_col=1, min_row=dataStartRow-1, max_col=endDataCol, max_row=excelSheet.max_row):
            # Stop Collecting Data When there is No More
            if dataRow[0].value == None:
                break
            
            # Get Data
            featuresTimesHolder[featureInd].append(float(dataRow[0].value))
            features[featureInd].append([float(dataRow[dataInd].value or 0) for dataInd in range(1, len(row))])
            
        return featuresTimesHolder, features, featureNames
    
    def getFeatures(self, featureOrder, inputFile = None, featureNames = None, surveyQuestions = [], finalSubjectInformationQuestions = []):
        # Load the Data from the Excel File
        xlWorkbook = xl.load_workbook(inputFile, data_only=True, read_only=True)
        worksheets = xlWorkbook.worksheets
        
        # Initialize experimental information
        experimentTimes = []; experimentNames = []
        surveyAnswerTimes = []; surveyAnswersList = [];
        # Initialize suject information
        subjectInformationAnswers = []; subjectInformationQuestions = []
        
        # Initialize data structures for feature parameters.
        featuresHolder = [[] for _ in range(len(featureOrder))]
        featuresTimesHolder = [[] for _ in range(len(featureOrder))]
        if featureNames == None:
            featureNames = [[] for _ in range(len(featureOrder))]
        
        # Loop through and compile all the data in the file
        for excelSheet in worksheets:
            # Extract experiment information
            if self.experimentalInfo_SheetName in excelSheet.title:
                experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions = self.extractExperimentalInfo(excelSheet, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions)
            # Extract subject information
            elif self.subjectInfo_SheetName in excelSheet.title:
                subjectInformationAnswers, subjectInformationQuestions = self.extractSubjectInfo(excelSheet, subjectInformationAnswers, subjectInformationQuestions)
            # Extract the features
            elif self.rawFeatures_AppendedSheetName in excelSheet.title:
                featuresTimesHolder, featuresHolder, featureNames = self.extractFeatures(excelSheet, featureOrder, featuresHolder, featuresTimesHolder, featureNames)
            else:
                sys.exit("Unsure what is in this file's excel sheet':", excelSheet.title)
    
        # Check that the subject background questions are all the same
        if len(finalSubjectInformationQuestions) != 0:
            assert all(np.array(finalSubjectInformationQuestions) == subjectInformationQuestions)
        
        return featuresTimesHolder, featuresHolder, featureNames, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions, subjectInformationAnswers, subjectInformationQuestions
    
    def streamTrainingData(self, trainingDataExcelFolder, featureOrder, numberOfChannels, readData, featureAverageIntervals = [],
                           indivisualFeatureNames = None, plotFeatures = False, reanalyzeData = False, blinkComparison = False):
        """
        Parameters
        ----------
        trainingDataExcelFolder: The Folder with ONLY the Training Data Excel Files
        """
        # Hold features from the data
        allRawFeatureHolders = []; allRawFeatureTimesHolders = []
        allAlignedFeatureHolder = []; allAlignedFeatureTimes = []
        analyzedFilesOrder = []
        # Hold survey information
        surveyQuestions = []; subjectInformationQuestions = []
        surveyAnswersList = []; surveyAnswerTimes = []
                        
        # For each file in the training folder
        for excelFile in list(natsorted(os.listdir(trainingDataExcelFolder))):
            # Get each excel file. This file should have the training data.
            if excelFile.endswith(".xlsx") and not excelFile.startswith("~"):
                trainingExcelFile = trainingDataExcelFolder + excelFile
                print("\nLoading Excel File", trainingExcelFile)
                excelFileName = excelFile.split(".")[0]
                analyzedFilesOrder.append(excelFileName)
                
                # Get the feature label from the filename
                # if blinkComparison:
                #     featureLabel = None
                #     for possibleFeatureLabel in ["Blink", "Movement", "Wire"]:
                #         if possibleFeatureLabel.lower() in excelFile.lower():
                #             featureLabel = possibleFeatureLabel
                #             break
                #     # If No Label, We Dont Want to Analyze its Features
                #     if featureLabel == None:
                #         sys.exit("No Feature Detected in File " + excelFile)
                #     surveyAnswersList.append(featureLabel)
                
                # ---------------------------------------------------------- #
                # -------------- Extract the Raw Training Data ------------- #
                
                savedFeaturesFile = trainingDataExcelFolder + self.saveFeatureFolder + excelFile.split(".")[0] + self.saveFeatureFile_Appended
                # If you want to and can use previously extracted features
                if not reanalyzeData and os.path.isfile(savedFeaturesFile):
                    rawFeatureTimesHolder, rawFeatureHolder, indivisualFeatureNames, experimentTimes, experimentNames, currentSurveyAnswerTimes, currentSurveyAnswersList, surveyQuestions, currentSubjectInformationAnswers, subjectInformationQuestions = self.getFeatures(featureOrder, inputFile = savedFeaturesFile, featureNames = indivisualFeatureNames, surveyQuestions = surveyQuestions, finalSubjectInformationQuestions = subjectInformationQuestions)
                else:
                    # Read in the training file with the raw data,
                    WB = xl.load_workbook(trainingExcelFile, data_only=True, read_only=True)
                    worksheets = WB.worksheets
                    
                    # Extract and analyze the raw data.
                    compiledRawData, experimentTimes, experimentNames, currentSurveyAnswerTimes, currentSurveyAnswersList, surveyQuestions, currentSubjectInformationAnswers, subjectInformationQuestions = self.extractExperimentalData(worksheets, numberOfChannels, surveyQuestions = surveyQuestions, finalSubjectInformationQuestions = subjectInformationQuestions)
                    ###### DELETE LATER!!!!
                    for axis in range(len(compiledRawData[1])):
                        compiledRawData[1][axis] = np.interp(np.arange(compiledRawData[0][0], compiledRawData[0][-1], 1/313), compiledRawData[0], compiledRawData[1][axis])
                    compiledRawData[0] = list(np.arange(compiledRawData[0][0], compiledRawData[0][-1], 1/313))
                    ################
                    readData.streamExcelData(compiledRawData, experimentTimes, experimentNames, currentSurveyAnswerTimes, currentSurveyAnswersList, surveyQuestions, currentSubjectInformationAnswers, subjectInformationQuestions, predictionModel = None, actionControl = None)          
                    
                    # Extract information from the streamed data
                    rawFeatureHolder = readData.rawFeatureHolder.copy()
                    rawFeatureTimesHolder = readData.rawFeatureTimesHolder.copy()
                    # Remove all previous information from this trial
                    readData.resetGlobalVariables()
                    
                    # Save the features to be analyzed in the future.
                    saveInputs = saveExcelData()
                    saveInputs.saveRawFeatures(rawFeatureTimesHolder, rawFeatureHolder, indivisualFeatureNames, featureOrder, experimentTimes, experimentNames, currentSurveyAnswerTimes,
                                               currentSurveyAnswersList, surveyQuestions, currentSubjectInformationAnswers, subjectInformationQuestions, trainingExcelFile)
                # Assert consistency across training data
                assert len(featureOrder) == len(rawFeatureHolder), "Incorrect number of channels"
                # Checkpoint: are there features in ALL categories
                for biomarkerInd in range(len(rawFeatureHolder)):
                    assert len(rawFeatureHolder[biomarkerInd]) > 1, "\tMissing raw features in " + featureOrder[biomarkerInd].upper() + " signal"
                
                # ---------------------------------------------------------- #
                # ------------------ Compile the Raw Data ------------------ #
                # Setup the compilation variables
                allRawFeatureTimesHolders.append(rawFeatureTimesHolder)
                allRawFeatureHolders.append(rawFeatureHolder); 
                featureInformation = [[], []]

                # Average the features across a sliding window
                for biomarkerInd in range(len(rawFeatureHolder)):
                    rawFeatures = np.array(rawFeatureHolder[biomarkerInd])
                    rawFeatureTimes = np.array(rawFeatureTimesHolder[biomarkerInd])
                                        
                    compiledFeatures = []
                    # Average the Feature Together at Each Point
                    for featureInd in range(len(rawFeatures)):
                        # Get the interval of features to average
                        featureMask = np.logical_and(
                            rawFeatureTimes <= rawFeatureTimes[featureInd],
                            rawFeatureTimes >= rawFeatureTimes[featureInd] - featureAverageIntervals[biomarkerInd]
                        )
                        featureInterval = rawFeatures[featureMask]
                        
                        # Take the trimmed average
                        compiledFeature = stats.trim_mean(featureInterval, 0.3)
                        compiledFeatures.append(compiledFeature)
                    # Store the compiled feature information
                    featureInformation[0].append(rawFeatureTimes)
                    featureInformation[1].append(compiledFeatures)
                                            
                # ---------------------------------------------------------- #
                # ------------- Align the Compiled Data in Time ------------ #
                minTimePoint = min(rawFeatureTimes[-1] if len(rawFeatureTimes) != 0 else 0 for rawFeatureTimes in featureInformation[0])
                readData.alignFeatures(lastTimePoint = minTimePoint, featureInformation = featureInformation)
                # Get the aligned feature information
                alignedFeatures = np.array(readData.alignedFeatures)
                alignedFeatureTimes = np.array(readData.alignedFeatureTimes)
                # Store the aligned feature information
                allAlignedFeatureHolder.append(alignedFeatures)
                allAlignedFeatureTimes.append(alignedFeatureTimes)
                
                # ---------------------------------------------------------- #
                # -------------------- Plot the features ------------------- #
                
                if plotFeatures:
                    # Initialize feature analysis class.
                    analyzeFeatures = featureAnalysis.featureAnalysis(trainingDataExcelFolder + "Data Analysis/")
                    featureNames = [item for sublist in indivisualFeatureNames for item in sublist]
                    
                    startAlignedIndex = 0
                    # For all the biomarkers in the experiment.
                    for biomarkerInd in range(len(rawFeatureHolder)):
                        # Get the training data's raw features information
                        rawFeatures = np.array(rawFeatureHolder[biomarkerInd])
                        rawFeatureTimes = np.array(rawFeatureTimesHolder[biomarkerInd])
                        # Get the training data's aligned features information
                        alignedFeatureSet = alignedFeatures[:, startAlignedIndex:startAlignedIndex + len(indivisualFeatureNames[biomarkerInd])]
                        startAlignedIndex += len(indivisualFeatureNames[biomarkerInd])
                        
                        # # Plot each biomarker's features from the training file.
                        analyzeFeatures.singleFeatureAnalysis(rawFeatureTimes, rawFeatures, indivisualFeatureNames[biomarkerInd], preAveragingSeconds = 0, averageIntervalList = [60, 120, 180], featureCollectionTimes = currentSurveyAnswerTimes, experimentTimes = experimentTimes, experimentNames = experimentNames, folderName = excelFileName + "/Feature Analysis/singleFeatureAnalysis - " + featureOrder[biomarkerInd].upper() + "/")
                        analyzeFeatures.singleFeatureAnalysis(alignedFeatureTimes, alignedFeatureSet, indivisualFeatureNames[biomarkerInd], preAveragingSeconds = featureAverageIntervals[biomarkerInd], averageIntervalList = [0], featureCollectionTimes = currentSurveyAnswerTimes, experimentTimes = experimentTimes, experimentNames = experimentNames, folderName = excelFileName + "/Feature Analysis/alignedFeatureAnalysis - " + featureOrder[biomarkerInd].upper() + "/")
                        # Plot the correlation across features
                        # analyzeFeatures.correlationMatrix(alignedFeatures, featureNames, folderName = "correlationMatrix/") # Hurts Plotting Style

                # ---------------------------------------------------------- #
                # ------------ Match Each Label with its Feature ----------- #
                
                # if blinkComparison:
                #     surveyAnswersList.extend(len(rawFeatureInfo) * [featureLabel])
                
                ### TODO: Sort the features with their correct labels.
                
                
                # Save the survey labels.
                surveyAnswersList.extend(readData.surveyAnswersList)         
                surveyAnswerTimes.append(readData.surveyAnswerTimes) 
                
                # ---------------------------------------------------------- #

        # Return Training Data and Labels
        return analyzedFilesOrder, allRawFeatureTimesHolders, allRawFeatureHolders, allAlignedFeatureTimes, allAlignedFeatureHolder, \
                np.array(surveyQuestions), np.array(surveyAnswersList), np.array(surveyAnswerTimes)
        
    def varyAnalysisParam(self, dataFile, featureOrder, numberOfChannels, readData, featureTimeWindows):
        print("\nLoading Excel File", dataFile)
        # Read in the training file with the raw data,
        WB = xl.load_workbook(dataFile, data_only=True, read_only=True)
        worksheets = WB.worksheets
        
        allRawFeatureTimesHolders = []; allRawFeatureHolders = []
        
        # Extract and analyze the raw data.
        compiledRawData, experimentTimes, experimentNames, currentSurveyAnswerTimes, currentSurveyAnswersList, surveyQuestions, currentSubjectInformationAnswers, subjectInformationQuestions = self.extractExperimentalData(worksheets, numberOfChannels, surveyQuestions = [], finalSubjectInformationQuestions = [])
        # For each test parameter
        for featureTimeWindow in featureTimeWindows:
            print(featureTimeWindow)
            # Set the parameter in the analysis
            readData.setFeatureWindowEEG(featureTimeWindow)
            
            # Stream the data
            readData.streamExcelData(compiledRawData, experimentTimes, experimentNames, currentSurveyAnswerTimes, currentSurveyAnswersList, surveyQuestions, currentSubjectInformationAnswers, subjectInformationQuestions, predictionModel = None, actionControl = None)          
            # Extract information from the streamed data
            allRawFeatureTimesHolders.append(readData.rawFeatureTimesHolder.copy())
            allRawFeatureHolders.append(readData.rawFeatureHolder.copy()); 
            # Remove all previous information from this trial
            readData.resetGlobalVariables()
            
            # Assert consistency across training data
            assert len(featureOrder) == len(allRawFeatureHolders[-1]), "Incorrect number of channels"
            # Checkpoint: are there features in ALL categories
            # for biomarkerInd in range(len(allRawFeatureHolders[-1])):
                # assert len(allRawFeatureHolders[-1][biomarkerInd]) > 1, "\tMissing raw features in " + featureOrder[biomarkerInd].upper() + " signal"
            
        return allRawFeatureTimesHolders, allRawFeatureHolders

# -------------------------------------------------------------------------- #
# -------------------------- Saving Data in Excel -------------------------- #

class saveExcelData(handlingExcelFormat):
    
    def getExcelDocument(self, excelFile, overwriteSave = False):
        # If the excel file you are saving already exists.
        if os.path.isfile(excelFile):
            # If You Want to Overwrite the Excel.
            if overwriteSave:
                print("\t\tDeleting Old Excel Workbook")
                os.remove(excelFile) 
            else:
                print("\t\tNot overwriting the file ... but your file already exists??")
            
        # If the File is Not Present: Create The Excel File
        if not os.path.isfile(excelFile):
            print("\t\tCreating New Excel Workbook")
            # Make Excel WorkBook
            WB = xl.Workbook()
            worksheet = WB.active 
            worksheet.title = self.emptySheetName
        else:
            print("\t\tExcel File Already Exists. Adding New Sheet to File")
            WB = xl.load_workbook(excelFile, read_only=False)
            worksheet = WB.create_sheet(self.emptySheetName)
        return WB, worksheet

    def addSubjectInfo(self, WB, worksheet, subjectInformationAnswers, subjectInformationQuestions):
        # Assert that the data is in the correct configuration
        assert len(subjectInformationAnswers) == len(subjectInformationQuestions)
        
        # Get the information ready for the file
        header = ["Background Questions", "Answers"]
        subjectInformationPointer = 0
                
        # Loop through/save all the data in batches of maxAddToexcelSheet.
        for firstIndexInFile in range(0, len(subjectInformationQuestions), self.maxAddToexcelSheet):
            # Add the information to the page
            worksheet.title = self.subjectInfo_SheetName
            worksheet.append(header)  # Add the header labels to this specific file.
            
            # Add the info to the first page
            while subjectInformationPointer != len(subjectInformationQuestions):
                # Add the data row to the worksheet
                row = [subjectInformationQuestions[subjectInformationPointer], subjectInformationAnswers[subjectInformationPointer]]
                worksheet.append(row)
                
                subjectInformationPointer += 1
                # Move onto next excel sheet if too much data
                if int(subjectInformationPointer/(firstIndexInFile+1)) == self.maxAddToexcelSheet:
                    break

            # Finalize document
            worksheet = self.addExcelAesthetics(worksheet) # Add Excel Aesthetics
            worksheet = WB.create_sheet(self.emptySheetName) # Add Sheet
        # Remove empty page
        WB.remove(worksheet)
    
    def addExperimentInfo(self, WB, worksheet, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions):
        # Assert that the data is in the correct configuration
        assert len(experimentTimes) == len(experimentNames)
        assert len(surveyAnswerTimes) == len(surveyAnswersList)
        # Set pointer
        experimentInfoPointer = 0; featureInfoPointer = 0
    
        # Get the Header for the experiment and survey
        header = ["Start Experiment (Seconds)", "End Experiment (Seconds)", "Experiment Label"]
        header.append("Feature Collection (Seconds)")
        header.extend(surveyQuestions)
                
        # Loop through/save all the data in batches of maxAddToexcelSheet.
        for firstIndexInFile in range(0, max(len(experimentTimes), len(surveyAnswerTimes)), self.maxAddToexcelSheet):
            # Add the information to the page
            worksheet.title = self.experimentalInfo_SheetName
            worksheet.append(header)  # Add the header labels to this specific file.
            
            # Add the info to the first page
            while experimentInfoPointer != len(experimentTimes) or featureInfoPointer != len(surveyAnswerTimes):
                row = []
                
                # Add experimental information
                if experimentInfoPointer != len(experimentTimes):
                    row.extend(experimentTimes[experimentInfoPointer])
                    row.append(experimentNames[experimentInfoPointer])
                    experimentInfoPointer += 1
                else:
                    row.extend([None]*3)
                # Add feature information
                if featureInfoPointer != len(surveyAnswerTimes):
                    row.append(surveyAnswerTimes[featureInfoPointer])
                    row.extend(surveyAnswersList[featureInfoPointer])
                    featureInfoPointer += 1
                
                # Add the data row to the worksheet
                worksheet.append(row)
                # Move onto next excel sheet if too much data
                if int(experimentInfoPointer/(firstIndexInFile+1)) == self.maxAddToexcelSheet or int(featureInfoPointer/(firstIndexInFile+1)) == self.maxAddToexcelSheet:
                    break

            # Finalize document
            worksheet = self.addExcelAesthetics(worksheet) # Add Excel Aesthetics
            worksheet = WB.create_sheet(self.emptySheetName) # Add Sheet
        # Remove empty page
        WB.remove(worksheet)
    
    def saveData(self, timePoints, signalData, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, 
                 surveyQuestions, subjectInformationAnswers, subjectInformationQuestions, dataHeaders, saveExcelPath):
        print("\n\tSaving raw signals")
        # ------------------------------------------------------------------ #
        # -------------------- Setup the excel document -------------------- #
        # Create the path to save the excel file.
        os.makedirs(os.path.dirname(saveExcelPath), exist_ok=True) # Create Output File Directory to Save Data: If None Exists
        
        # Get the excel document.
        WB, worksheet = self.getExcelDocument(saveExcelPath, overwriteSave = False)

        # ------------------------------------------------------------------ #
        # -------------- Add experimental/subject information -------------- #
        # Add subject information
        self.addSubjectInfo(WB, worksheet, subjectInformationAnswers, subjectInformationQuestions)
        worksheet = WB.create_sheet(self.emptySheetName) # Add 
        # Add experimental information
        self.addExperimentInfo(WB, worksheet, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions)
        worksheet = WB.create_sheet(self.emptySheetName) # Add Sheet
        
        # ------------------------------------------------------------------ #
        # ---------------------- Add data to document ---------------------- #     
        # Get the Header for the Data
        header = ["Time (Seconds)"]
        header.extend([dataHeader.upper() + " Raw Data" for dataHeader in dataHeaders])
        
        # Loop through/save all the data in batches of maxAddToexcelSheet.
        for firstIndexInFile in range(0, len(timePoints), self.maxAddToexcelSheet):
            startTimer = timer.time()
            # Add the information to the page
            worksheet.title = self.rawSignals_Sheetname
            worksheet.append(header)  # Add the header labels to this specific file.
                        
            # Loop through all data to be saved within this sheet in the excel file.
            for dataInd in range(firstIndexInFile, min(firstIndexInFile+self.maxAddToexcelSheet, len(timePoints))):
                # Organize all the data
                row = [timePoints[dataInd]]
                row.extend([dataCol[dataInd] for dataCol in signalData])
                
                # Add the row to the worksheet
                worksheet.append(row)
    
            # Finalize document
            worksheet = self.addExcelAesthetics(worksheet) # Add Excel Aesthetics
            worksheet = WB.create_sheet(self.emptySheetName) # Add Sheet
            
            # If I need to use another sheet
            if firstIndexInFile + self.maxAddToexcelSheet < len(timePoints):
                # Keep track of how long it is taking.
                endTimer = timer.time()
                numberOfSheetsLeft = 1+(len(timePoints) - firstIndexInFile - self.maxAddToexcelSheet)//self.maxAddToexcelSheet
                timeRemaining = (endTimer - startTimer)*numberOfSheetsLeft
                print("\tEstimated Time Remaining " + str(timeRemaining) + " seconds; Excel Sheets Left to Add: " + str(numberOfSheetsLeft))
        # Remove empty page
        if worksheet.title == self.emptySheetName:
            WB.remove(worksheet)
        
        # ------------------------------------------------------------------ #
        # ------------------------ Save the document ----------------------- #  
        # Save as New Excel File
        WB.save(saveExcelPath)
        WB.close()
            
    def saveRawFeatures(self, rawFeatureTimesHolder, rawFeatureHolder, indivisualFeatureNames, featureOrder, experimentTimes, experimentNames,
                     surveyAnswerTimes, surveyAnswersList, surveyQuestions, subjectInformationAnswers, subjectInformationQuestions, excelFilename): 
        print("\n\tSaving raw features")
        # ------------------------------------------------------------------ #
        # -------------------- Setup the excel document -------------------- #
        # Organize variables
        baseFilename = os.path.basename(excelFilename).split('.')[0]
        excelPath = excelFilename.split(baseFilename)[0]
        
        # Create the path to save the excel file.
        saveDataFolder = excelPath + self.saveFeatureFolder
        os.makedirs(saveDataFolder, exist_ok=True) # Create Output File Directory to Save Data: If None Exists
        # Specify the name of the file to save
        saveExcelName = baseFilename + self.saveFeatureFile_Appended
        excelFile = saveDataFolder + saveExcelName
        
        # Get the excel document.
        WB, worksheet = self.getExcelDocument(excelFile, overwriteSave = True)
        
        # ------------------------------------------------------------------ #
        # -------------- Add experimental/subject information -------------- #
        # Add subject information
        self.addSubjectInfo(WB, worksheet, subjectInformationAnswers, subjectInformationQuestions)
        worksheet = WB.create_sheet(self.emptySheetName) # Add 
        # Add experimental information
        self.addExperimentInfo(WB, worksheet, experimentTimes, experimentNames, surveyAnswerTimes, surveyAnswersList, surveyQuestions)
        worksheet = WB.create_sheet(self.emptySheetName) # Add Sheet

        # ------------------------------------------------------------------ #
        # ---------------------- Add data to document ---------------------- #  
        # Indivisually add features from each sensor to the excel file.
        for featureTypeInd in range(len(featureOrder)):
            featureNames = indivisualFeatureNames[featureTypeInd]
            # Extract raw features
            featureTimes = rawFeatureTimesHolder[featureTypeInd]
            rawFeatures = rawFeatureHolder[featureTypeInd] 
            
            # Create the header bar
            header = ["Time (Seconds)"]
            header.extend(featureNames)
        
            # Loop through/save all the data in batches of maxAddToexcelSheet.
            for firstIndexInFile in range(0, len(featureTimes), self.maxAddToexcelSheet):
                # Add the information to the page
                worksheet.title = featureOrder[featureTypeInd].upper() + self.rawFeatures_AppendedSheetName # Add the sheet name to the file
                worksheet.append(header)  # Add the header labels to this specific file.
                            
                # Loop through all data to be saved within this sheet in the excel file.
                for dataInd in range(firstIndexInFile, min(firstIndexInFile+self.maxAddToexcelSheet, len(featureTimes))):
                    # Organize all the data
                    row = [featureTimes[dataInd]]
                    row.extend(rawFeatures[dataInd])
                                    
                    # Add the row to the worksheet
                    worksheet.append(row)
        
                # Finalize document
                worksheet = self.addExcelAesthetics(worksheet) # Add Excel Aesthetics
                worksheet = WB.create_sheet(self.emptySheetName) # Add Sheet
        # Remove empty page
        if worksheet.title == self.emptySheetName:
            WB.remove(worksheet)
            
        # ------------------------------------------------------------------ #
        # ------------------------ Save the document ----------------------- #  
        # Save as New Excel File
        WB.save(excelFile)
        WB.close()
